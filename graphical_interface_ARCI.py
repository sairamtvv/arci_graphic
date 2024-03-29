import tkinter as TK
from tkinter import filedialog
from tkinter import messagebox,StringVar,ttk

import time   #for time.sleep
import sys
import pathlib
import os, stat
import shutil

import numpy as np
import pandas as pd
import openpyxl


import socket
from SCPI_socket import SCPI_sock_connect,SCPI_sock_send,SCPI_sock_close,getDataFromSocket
from matlabconversionprog import matlab_conver_func
from selenium import webdriver
from selenium.webdriver.common.keys import Keys



import pywinauto
from pywinauto.application import Application
import sys
from pywinauto.findwindows import WindowAmbiguousError, WindowNotFoundError
from pywinauto.controls.common_controls import TabControlWrapper
from pywinauto.keyboard import send_keys, KeySequenceError






#import autoit

#autoit.opt("MouseCoordMode", 0)
#autoit.opt("SendKeyDelay", 10)

"""
autoit.run("C:\Program Files (x86)\Lumenera Corporation\LuCam Capture Software\LuCam.exe")
autoit.win_wait('LuCam Capture')
autoit.win_activate('LuCam Capture')

time.sleep(200/1000)  #time is in seconds

autoit.mouse_click('primary', 54, 169, 1, 0)  #Start preview
time.sleep(200/1000)
#Send ('{TAB}{TAB}{Enter}')


autoit.win_wait_active("[CLASS:Notepad]", 3)
autoit.control_send("[CLASS:Notepad]", "Edit1", "hello world{!}")
autoit.win_close("[CLASS:Notepad]")
autoit.control_click("[Class:#32770]", "Button2")
"""

class Window():
    
    root_folder="D:/Automation_base/"
    #Location where resources can be found
    resourcelocation=root_folder+'resources/'

    def __init__(self,master):

        self.master=master
        master.title("DATA ACQUISITION SYSTEM ARCI")



        #Standard Locations
        #self.baselocation='D:/'
        #The default location from which the file shall be read from
        self.defaultlocation=Window.root_folder
        #self.resourcelocation=self.baselocation+'resources/'















        # Given in minutes
        self.intradelaydefault='60'
        self.largedelaydefault=str(24*60)
        self.sensordefault='255'
        #temprunsdefault is for L1,L2 and L3
        self.temprunsdefault='9'
        #temprunsdefault45  is for Day 4 and Day 5
        self.temprunsdefault45='3'





        # Button for opening a file
        self.buttonopenfile = TK.Button(text=self.defaultlocation, command=self.openfile, width=50)
        self.buttonopenfile.grid(row=2, column=0)
        self.labelopenfile=TK.Label(text="File location that is read", fg='red')
        self.labelopenfile.grid(row=1, column=0)


        #White Space area
        self.output = TK.Text()
        self.output.grid(row=7,column=0)

        #Progress bar
        self.progressbar= ttk.Progressbar(length=150)
        self.progressbar.grid(row=5,column=0)
        self.progressbar.config(maximum=7, value=0)
        #Label for Progress Bar
        self.labelprogress=TK.Label(text="Progressbar", fg='red')
        self.labelprogress.grid(row=4, column=0)


        #Delay will be entry boxes due to its float nature
        #intradelay
        intravar=StringVar(value=self.intradelaydefault)
        self.intraentry=TK.Entry(textvariable=intravar)
        self.intraentry.grid(row=3,column=1)
        #label for intradelay
        self.labelintraentry=TK.Label(text="INTRA DELAY (mins)", fg='red')
        self.labelintraentry.grid(row=2,column=1)




        #Sensor Number
        sensorvar = StringVar(value=self.sensordefault)
        self.sensorentry=TK.Entry(textvariable=sensorvar)
        self.sensorentry.grid(row=2,column=2)
        #label for Sensor Number
        self.labelsensornumber=TK.Label(text="SENSOR NUMBER", fg='red')
        self.labelsensornumber.grid(row=1,column=2)

        #Number of Temperature runs for L1, L2 and L3
        temprunsvar=StringVar(value=self.temprunsdefault)  # temporary variable
        self.temprunsentry=TK.Entry(textvariable=temprunsvar)
        self.temprunsentry.grid(row=5,column=2)
        #Label for number of temperature runs
        self.labeltempruns=TK.Label(text="Number of Temperature Runs", fg='red')
        self.labeltempruns.grid(row=4,column=2)


        #Number of Temperature runs for D4 and D5
        temprunsvar_D45=StringVar(value=self.temprunsdefault45)  # temporary variable
        self.temprunsentry_D45=TK.Entry(textvariable=temprunsvar_D45)
        self.temprunsentry_D45.grid(row=5,column=1)
        #Label for number of temperature runs
        self.labeltempruns_D45=TK.Label(text="Number of Temperature Runs", fg='red')
        self.labeltempruns_D45.grid(row=4,column=1)



        #Drop down list for the days
        self.combo = ttk.Combobox()
        self.combo.grid(row=1, column=3)
        self.combo.config(value = ('L1', 'L2', 'L3', 'D4', 'D5','D6','set values'))
        self.combo.set('L3')
        #Label for combo box


        
        
        
        #Button for Validation
        self.buttonvalidate = TK.Button(text='Validate', command=self.validate)
        self.buttonvalidate.grid(row=3, column=3)
        self.labelupdate=TK.Label(text="Validation of your values", fg='red')
        #self.labelopenfile.grid(row=1, column=3)

        #Button to specify the next run in how much time


        # Button for Analysis
        self.buttonanalysis = TK.Button(text='Analyse', command=self.analysis_alldays)
        self.buttonanalysis.grid(row=6, column=3)
        

       #Drop down for internal or external files for Analysis 
       #If internal the location is read from the previously run code
       #else the location is read from the external 
        self.analysiscombo = ttk.Combobox()
        self.analysiscombo.grid(row=7, column=3)
        self.analysiscombo.config(value = ('Internal', 'External'))
        self.analysiscombo.set('Internal')











    def validate(self):


        #delay after a click
        self.general_time_delay=200/1000 # in milliseconds
        #delay after a winwait etc.
        self.more_time_delay=500/1000 # in milliseconds





        self.intradelay_set=self.intraentry.get()
        print("Delay is {}".format(self.intradelay_set))
        self.sensor_input_set=self.sensorentry.get()
        print("Sensor Number is {}".format(self.sensor_input_set))
        self.tempruns_set=self.temprunsentry.get()
        print("Number of temperature runs for L1,L2 and L3 are {}".format(self.tempruns_set))
        self.comboday_set=self.combo.get()
        print("The day you want to run is {}".format(self.comboday_set))
        self.tempruns_D45_set=self.temprunsentry_D45.get()
        print("Number of temperature runs for Day4 and Day5 are {}".format(self.tempruns_D45_set))
        print("Validating the Path and existence of other files::")
        self.analysiscombo_set=self.analysiscombo.get()
        print("The type of analysis you want is ".format(self.analysiscombo_set))



        # All Paths  needs to be here as sensorname is part of the path
        #Base bolder where all the folders shall be made
        self.root_folder="D:/Automation_base/"
        self.base_folder = "D:/Automation_base/" + self.sensor_input_set +'/'
        self.base_folder_path= pathlib.Path(self.base_folder)

        #default location where the datalogger saves the file for each scan
        self.dataloggerlocation="C:/Users/PRASAD/Documents/"
        #default path where the datalogger saves the file for each scan
        self.path_to_datalogger_default=pathlib.Path(self.dataloggerlocation)

        

        #Location where the Analysis is done
        self.analysislocation=self.base_folder+'analysis_folder/'

        #raw text or output  raw_text.txt file location
        self.raw_text_path=pathlib.Path(self.base_folder+'raw_text_'+self.comboday_set+'.txt')
        self.combolocation=self.base_folder + self.comboday_set +'/'












        if self.comboday_set == 'L1':

            self.createdirectory(self.base_folder_path)



        elif self.comboday_set == 'L2':
            self.checkfileexists(self.base_folder+'raw_text_L1.txt')


        elif self.comboday_set == 'L3':
            self.checkfileexists(self.base_folder+'raw_text_L1.txt')
            self.checkfileexists(self.base_folder+'raw_text_L2.txt')


        elif self.comboday_set == 'D4':
            self.checkfileexists(self.base_folder+'raw_text_L1.txt')
            self.checkfileexists(self.base_folder+'raw_text_L2.txt')
            self.checkfileexists(self.base_folder+'raw_text_L3.txt')

        elif (self.comboday_set == 'D5' and self.tempruns_D45_set==3):
            self.checkfileexists(self.base_folder+'D4summary_temp1.txt')
            self.checkfileexists(self.base_folder+'D4summary_temp2.txt')
            self.checkfileexists(self.base_folder+'D4summary_temp3.txt')
            self.checkfileexists(self.base_folder+'raw_text_D4_1_33.txt')
            self.checkfileexists(self.base_folder+'raw_text_D4_2_33.txt')
            self.checkfileexists(self.base_folder+'raw_text_D4_3_33.txt')

        else:
            print("\n")


        #espec details i.e temperature controller
        self.espec_HOST="192.168.1.18"
        self.espec_PORT=57732
        self.espec_session=SCPI_sock_connect(self.espec_HOST,self.espec_PORT)
        #print(self.espec_session)

        #Validating ESPEC temperature controller by reading its temperature once
        SCPI_sock_send(self.session,'TEMP?')
        output=self.session.recv(20).decode()
        lstoutput=output.split(",")

        #21.3,-40.0,165.0,-70
        if -110<float(lstoutput[0])<110:
            print("Espec Validated and could read temperature")
        else:
            print("Please Check your connection to ESPEC")
            self.master.destroy()
            sys.exit(0)



        #Validating the Aerotech Position Controller
        self.aerotech_url="192.168.1.16"
        #Install the latest Chrome
        path_to_chromedriver=pathlib.Path(self.resourcelocation+'chromedriver_win32/chromedriver.exe')
        self.driver=webdriver.Chrome(path_to_chromedriver)
        print("Waiting for the AEROTECH URL....")
        self.driver.get(self.aerotech_url)
        time.sleep(10)

        #Checking if the two_balls areon the screen
        balls_Element=self.driver.find_element_by_id("enableDisableAxis0")

        if balls_Element is None:
            print("The position controller is not connecting...")
            self.master.destroy()
            sys.exit(0)
        else:
            print("Connected to Aerotech controller...")



        #checking if Datalogger is active
        app = Application(backend="win32").connect(title=u'Configuration - 2 - BenchLink Data Logger 3', class_name='WindowsForms10.Window.8.app.0.33c0d9d')
        main_dlg = app[u'WindowsForms10.Window.8.app.0.33c0d9d']
        main_dlg.wait('visible')
        print("1")
        time.sleep(1)
        main_dlg.set_focus()
        print("2")
       
        p=main_dlg.TabControl.select(u'Scan and Log Data')
        #print(p.get_properties())

       #checking if dataloger default exists in the location specified.
       #The  file into which data is printed always
        #self.checkfileexists(self.dataloggerlocation)

        self.progressbar.config(value=int(self.comboday_set[1]))
        print("Validation succcessful please press run..\n")



    def checkfileexists(self, checkfilename):
        path_checkfilename=pathlib.Path(checkfilename)
        if not path_checkfilename.exists():
            print("Oops"+ checkfilename +" file doesn't exist!\n")
            self.master.destroy()
            sys.exit(0)
        else:
            print(checkfilename)
            return True

    def checkactivewindows(self,title):
        try:
            autoit.win_activate(title)
        except NameError:
            print ("The" + title + "is not active\n please open it\n")

        else:
            print ("Could Successfully activate" + title + "\n ")


    def closing_protocol(self):
        SCPI_sock_close(self.session)

    def openfile(self):
        print(messagebox.askquestion(title='Opening file', message='Do you want open file ?'))
        filechosen = filedialog.askopenfile()
        print(filechosen.name)
        self.defaultlocation=filechosen.name
        #defaultlocation
        #self.filename=StringVar(value=filechosen.name)
        print(self.defaultlocation)
        self.buttonopenfile.configure(text=self.defaultlocation)
        #self.entry = TK.Entry(textvariable=self.filename)
        self.output.insert(TK.END,self.defaultlocation)

    def analysis_russian(self):
        
        #list of all raw text files from all the days
        lstrawtext=['raw_text_L1.txt','raw_text_L2.txt','raw_text_L3.txt']
        #list of all the output files it needs to make
        lst_treatg_input=[self.sensor_input_set+"_AS_L1",self.sensor_input_set+"_AS_L2",self.sensor_input_set+"_AS_L3"]
        lst_channels101_to_106=["101 (VDC)","201 (VDC)","102 (VDC)","202 (VDC)","103 (VDC)","203 (VDC)","104 (VDC)","204 (VDC)","105 (VDC)","205 (VDC)","106 (VDC)","206 (VDC)","107 (VDC)","207 (VDC)","108 (VDC)","208 (VDC)"]
        lst_channels108_to_116=["109 (VDC)","209 (VDC)","110 (VDC)","210 (VDC)","111 (VDC)","211 (VDC)","112 (VDC)","212 (VDC)","113 (VDC)","213 (VDC)","114 (VDC)","214 (VDC)","115 (VDC)","215 (VDC)","116 (VDC)","216 (VDC)"]
        #list of all the channels
        lstchannels=lst_channels101_to_106+lst_channels108_to_116
        
        print("Preparing files for the TreatG software.. This takes a minute...\n")
        #knowing the number of channels from the column number of 201
        #Based on that the number of channels can be infered
        #Now only these many number of channel files shall be created 
        dfL1= pd.read_csv('raw_text_L1.txt')
        location201=dfL1.columns.get_loc("201 (VDC)")
        channels=int((location201-2)/2)
        #Writing heading to all the files L1, L2,L3 with all the channel numbers so that 
        #nextpart can append these files.The files will be L1_1 L1_2 till L1_channelnumber L2_channelnumber
        #L3_channelnumber
        for inputtreatg in lst_treatg_input:
            for channelnum in range(0,channels):
                with open(inputtreatg+"_"+str(channelnum+1), 'w') as f:
                    f.write(' pos  uacc     Rts\n')
        #This is the main part of the program
                    
         
        for rawtext_day,inputtreatg in zip(lstrawtext,lst_treatg_input):
            dfL1= pd.read_csv(rawtext_day)
            #This for loop chooses the 4 rows, then the interchange among them happens in the desired order
            #order from 1423 to 1234
            for chose4rowspanda in range(0,33,4):
                dfL1temp=dfL1.loc[chose4rowspanda:chose4rowspanda+3,:]
                b,c,d=dfL1temp.iloc[1,:].copy(),dfL1temp.iloc[2,:].copy(),dfL1temp.iloc[3,:].copy()
                dfL1temp.iloc[1,:],dfL1temp.iloc[2,:],dfL1temp.iloc[3,:]=c,d,b
                #for every channel selected, we have to assign the correct column names
                #It is like selecting a matrix using rows and columns channel number  gets the column name from the list
                for channelnum in range(0,channels):
                    with open(inputtreatg+"_"+str(channelnum+1), 'a') as f:
                        for rows in range(0,4):
                            f.write('  {0} {1:.6f}  {2:.5f}\n'.format(rows,dfL1temp[lstchannels[2*channelnum]].iloc[rows],dfL1temp[lstchannels[2*channelnum+1]].iloc[rows]))
        print("Finished Preparing files for treatg...\n")
        
    def analysis_alldays(self):
        
        #INTERNAL when the file location can be got from the runs
        if self.analysiscombo.get() == "Internal":
            dir_for_analy=self.analysislocation
            sensorname=self.sensor_input_set
            
            
            
            
            
        else:
            chosenfile=pathlib.Path(self.defaultlocation)
            chosenfile=pathlib.Path(self.defaultlocation)
                
            dir_for_analy=chosenfile.parent
            sensorname=input("Enter the sensor name:")
       
        
        
        
        
        
        lstresult_files=['RESULT1.DAT','RESULT2.DAT','RESULT3.DAT']
        lstfinalD45files=['finalD45_temp1.txt','finalD45_temp2.txt','finalD45_temp3.txt']        
        lstD4temp_33files=['raw_text_D4_1_33.txt','raw_text_D4_2_33.txt','raw_text_D4_3_33.txt']
        listof_filesrequired=lstresult_files+lstfinalD45files+lstD4temp_33files
        #printing all the required file names 
        for item in listof_filesrequired:
            print(item) 
            
        decision=input ("If files with all these names are present in folder \n  press Y else add these files:")
       
        if decision.lower() != "y":
            self.master.destroy()
            sys.exit(0)
          
        #For checking if all the required files are present, if not the program exits
        for item in listof_filesrequired:
            self.checkfileexists(dir_for_analy.joinpath(item))
        
        
            
        lst_of_excelfiles=['(AS)_adj_param_m40.xlsx','(AS)_adj_param_p70.xlsx','(AS)_adj_param_p20.xlsx']
        
        
        #g_sheet=srcfile.sheetnames
        
        #sheetname = srcfile.get_sheet_by_name('sheetsai')#get sheetname from the file
        #sheetname['C4']= 55.568 #write something in B2 cell of the supplied sheet
        
        lst_ten_numbers=list(range(0,10))
        
        
        #For all the files in the list writing 33 values in to the desired excel cells
        for  excelfile, file33 in zip(lst_of_excelfiles,lstD4temp_33files):
            for num in lst_ten_numbers:
                
                
                df_temp_33=pd.read_csv(dir_for_analy.joinpath(file33),header=None)
                
                srcfile = openpyxl.load_workbook(dir_for_analy.joinpath(excelfile),read_only=False, keep_vba= True)#to open the excel sheet and if it has macros
                sheetname=srcfile["sheetsai"]
                sheetname.cell(row=item+4,column=3).value = df_temp_33.loc[item,0] #write to row 1,col 1 explicitly, this type of writing is useful to write something in loops
            srcfile.save(dir_for_analy.joinpath(excelfile.split(".")[0]+"_modified"+".xlsm"))#save it as a new file, the original file is untouched and here I am saving it as xlsm(m here denotes macros).
             
        lst_of_xlsmfiles_modified=['(AS)_adj_param_m40_modified.xlsm','(AS)_adj_param_p70_modified.xlsm','(AS)_adj_param_p20_modified.xlsm']   
        lstsix_numbers=list(range(0,6))
        
        #For all the files in the list writing D45 values in to the desired excel cells
        for  xlsmfile,fileD45 in zip(lst_of_xlsmfiles_modified,lstD4temp_33files):
            for rows in lstsix_numbers:
                for columns in lstsix_numbers:
                
                    
                    df_temp_45=pd.read_csv(dir_for_analy.joinpath(fileD45),header=None)
                    
                    srcfile = openpyxl.load_workbook(dir_for_analy.joinpath(xlsmfile),read_only=False, keep_vba= True)#to open the excel sheet and if it has macros
                    sheetname=srcfile["sheetsai"]
                    sheetname.cell(row=rows+19,column=columns+2).value = df_temp_45.loc[rows,columns] #write to row 1,col 1 explicitly, this type of writing is useful to write something in loops
            srcfile.save(dir_for_analy.joinpath(sensorname+xlsmfile))
            
        #Now that the desired excel files are written deleting the extra unnecessary files 
        
        for file1,file2 in zip(lst_of_excelfiles,lst_of_xlsmfiles_modified):
                file_to_rem = dir_for_analy.joinpath(file1)
                file_to_rem.unlink()
                file_to_rem = dir_for_analy.joinpath(file2)
                file_to_rem.unlink()
        
        #Makng the desired array as the input for matlab conversion program
        df_for_matlab=df_temp_33=pd.read_csv(dir_for_analy.joinpath("raw_text_D4_1_33.txt"),header=None)
        desired_matrix=df_for_matlab.iloc[:,0:2].values
        
        
        #Calling the matlab conversion program
        matlab_conver_func(desired_matrix,dir_for_analy,sensorname)
        
        
        
        
        
        
        
        


    def acquistionL123(self):

        lsttemp=[70,42.5,15,-12.5,-40,-12.5,15,42.5,70]
        #range starts from 0
        #for tempiter in range(1,int(self.tempruns_set)+1):
        for tempiter,actual_temp in enumerate(lsttemp):

            tempiter=tempiter+1
            self.check_temp_reached(actual_temp)

            #---position one start--------
            #position numbers are importsnt as desired aerotech position depends on this number
            positionnumber=1
            self.desired_aerotech_pos(positionnumber)
            self.autoitdataloggerL123(tempiter,positionnumber)
            time.delay(self.intradelay_set*60)
            #-------finished one position------------------------------------------

            positionnumber=2
            self.desired_aerotech_pos(positionnumber)
            self.autoitdataloggerL123(tempiter,positionnumber)
            time.delay(self.intradelay_set*60)
            #-------finished two position------------------------------------------

            positionnumber=3
            self.desired_aerotech_pos(positionnumber)
            self.autoitdataloggerL123(tempiter,positionnumber)
            time.delay(self.intradelay_set*60)

            #-------finished THREE position------------------------------------------

            positionnumber=4
            self.desired_aerotech_pos(positionnumber)
            self.autoitdataloggerL123(tempiter,positionnumber)
            time.delay(self.intradelay_set*60)


    def createdirectory(self,location):
        if not os.path.exists(location):
            try:
                os.mkdir(location)
            except OSError:
                print ("Creation of the directory locaton %s failed " % location)
                self.master.destroy()
                sys.exit(0)
            else:
                print ("Successfully created  directory %s " %location)


    def check_temp_reached(self,desired_temp):
        SCPI_sock_send(self.session,'TEMP?')
        output=self.session.recv(20).decode()
        lstoutput=output.split(",")
        #21.3,-40.0,165.0,-70
        print(f"Waiting for temperature: {desired_temp}")
        while(float(lstoutput[0])!=desired_temp):
            time.sleep(60)
        print(f"Reached the desired temperature:{desired_temp}")




    def acquistionD4(self):
        #self.tempruns_D45_set  is number of temperature runs set for day 4 and day5
        lsttemp=[-40,70,20]
        #for tempiter in range(1,int(self.tempruns_D45)+1):
        for tempiter,actual_temp in enumerate(lsttemp):

            tempiter=tempiter+1
            self.check_temp_reached(actual_temp)
            #step 1 waiting for temperature ramp for 1 hr
            time.delay(60*60)
            #step 2 waiting at that temperature for 1 hr
            time.delay(60*60)


            #step 3 Acquistion by altering the position
            for cycle in range(1,7):
            # when ever  cycle is zero it is going to create a new file in autoitdatalogger45

                positionnumber =1
                self.desired_aerotech_pos(positionnumber)
                self.autoitdataloggerD45(self,tempiter,positionnumber,cycle)

                positionnumber =2
                self.desired_aerotech_pos(positionnumber)
                self.autoitdataloggerD45(self,tempiter,positionnumber,cycle)

            #step 4 waiting for 30 minutes
            time.delay(30*60)

            #step 5 Acquistion by altering the position 3 and 4
            for cycle in range(1,7):
            # when ever  cycle is zero it is going to create a new file in autoitdatalogger45

                positionnumber =3
                self.desired_aerotech_pos(positionnumber)
                self.autoitdataloggerD45(self,tempiter,positionnumber,cycle)

                positionnumber =4
                self.desired_aerotech_pos(positionnumber)
                self.autoitdataloggerD45(self,tempiter,positionnumber,cycle)


            #step 6 waiting for 30 minutes
            time.delay(30*60)


            #step 7 Acquiring 10 readings in position
            #temp_33 shall be the file name where the readings have been taken for 10 times
            positionnumber =33

            self.desired_aerotech_pos(3)

            for cycle in range(1,11):
                #print ("Start : %s" % time.ctime())
                self.autoitdataloggerD45(self,tempiter,positionnumber,cycle)
                #print ("Stop : %s" % time.ctime())

        self.summaryD4()



    def finalanalysisD45(self):
        dfD4summary_temp1=pd.read_csv('D4summary_temp1.txt')
        dfD5_temp1_pos5= pd.read_csv('raw_text_D5_1_5.txt')
        dfD5_temp1_pos6= pd.read_csv('raw_text_D5_1_6.txt')

        df_finalD45_temp1=pd.concat([dfD4summary_temp1,dfD5_temp1_pos5,dfD5_temp1_pos6],axis=1)
        df_finalD45_temp1.to_csv('finalD45_temp1.txt',index=False)


        dfD4summary_temp2=pd.read_csv('D4summary_temp2.txt')
        dfD5_temp2_pos5= pd.read_csv('raw_text_D5_2_5.txt')
        dfD5_temp2_pos6= pd.read_csv('raw_text_D5_2_6.txt')

        df_finalD45_temp2=pd.concat([dfD4summary_temp2,dfD5_temp2_pos5,dfD5_temp2_pos6],axis=1)
        df_finalD45_temp2.to_csv('finalD45_temp2.txt',index=False)

        dfD4summary_temp3=pd.read_csv('D4summary_temp3.txt')
        dfD5_temp3_pos5= pd.read_csv('raw_text_D5_3_5.txt')
        dfD5_temp3_pos6= pd.read_csv('raw_text_D5_3_6.txt')

        df_finalD45_temp3=pd.concat([dfD4summary_temp3,dfD5_temp3_pos5,dfD5_temp3_pos6],axis=1)
        df_finalD45_temp3.to_csv('finalD45_temp3.txt',index=False)
        
        
        dfD4_temp1_pos33=pd.read_csv('raw_text_D4_1_33.txt')
        dfD4_temp2_pos33=pd.read_csv('raw_text_D4_2_33.txt')
        dfD4_temp3_pos33=pd.read_csv('raw_text_D4_3_33.txt')




    def summaryD4(self):
        "this Concat does along the columns"

        if self.tempruns_D45==3:

            dfD4_temp1_pos1= pd.read_csv('raw_text_D4_1_1.txt')
            dfD4_temp1_pos2= pd.read_csv('raw_text_D4_1_2.txt')
            dfD4_temp1_pos3= pd.read_csv('raw_text_D4_1_3.txt')
            dfD4_temp1_pos4= pd.read_csv('raw_text_D4_1_4.txt')

            D4summary_temp1=pd.concat([dfD4_temp1_pos1,dfD4_temp1_pos2,dfD4_temp1_pos3,dfD4_temp1_pos4],axis=1)
            D4summary_temp1.to_csv('D4summary_temp1.txt',index=False)

            dfD4_temp2_pos1= pd.read_csv('raw_text_D4_2_1.txt')
            dfD4_temp2_pos2= pd.read_csv('raw_text_D4_2_2.txt')
            dfD4_temp2_pos3= pd.read_csv('raw_text_D4_2_3.txt')
            dfD4_temp2_pos4= pd.read_csv('raw_text_D4_2_4.txt')

            D4summary_temp2=pd.concat([dfD4_temp2_pos1,dfD4_temp2_pos2,dfD4_temp2_pos3,dfD4_temp2_pos4],axis=1)
            D4summary_temp2.to_csv('D4summary_temp2.txt',index=False)

            dfD4_temp3_pos1= pd.read_csv('raw_text_D4_3_1.txt')
            dfD4_temp3_pos2= pd.read_csv('raw_text_D4_3_2.txt')
            dfD4_temp3_pos3= pd.read_csv('raw_text_D4_3_3.txt')
            dfD4_temp3_pos4= pd.read_csv('raw_text_D4_3_4.txt')

            D4summary_temp3=pd.concat([dfD4_temp3_pos1,dfD4_temp3_pos2,dfD4_temp3_pos3,dfD4_temp3_pos4],axis=1)
            D4summary_temp3.to_csv('D4summary_temp3.txt',index=False)

    def autoitdataloggerD45(self,tempiter,positionnumber,cycle):

        self.autoit_singlecycleclicks()

        """
        The methodology used here is for every new temperature and position file is created.
        temp_position 1_1 to 1_4 , 2_1 to 2_4, 3_1 to 3_4
        in cycle 1 the file is created as it is in write mode. When ever heading is printed it is going
        to create file. The last cycle shall be in 33 file.1_33,2_33,raw_text_D4_3_33
        total 15 files shall be outputon day 4.



        """
        files = []
        # r=root, d=directories, f = files
        for r, d, f in os.walk(self.path_to_datalogger_default):
            for file in f:
                files.append(os.path.join(r, file))


        #Assuming it has saved in dataloggerdefault_file.CSV
        #Name to which each of the cycles file shall be saved
        self.D45_path=pathlib.Path(self.base_folder+'raw_text_'+self.comboday_set+'_'+tempiter +'_'+ positionnumber+'.txt')
        with open(files[0],encoding='UTF-16') as f:
            df_datalogger_default = pd.read_csv(f,skiprows=20,header=None)
            if (cycle == 1):
                df_datalogger_default.loc[[0]].to_csv(self.D45_path,index=False, header=False )  #heading
            df_datalogger_default.loc[[1]].to_csv(self.D45_path,index=False, header=False, mode='a')


        my_file=files[0]
        to_file=pathlib.Path(self.combolocation + tempiter + '_' + positionnumber +'.txt')
        shutil.copy(my_file, to_file)


    def acquistionD5(self):
        #self.tempruns_D45_set  is number of temperature runs set for day 4 and day5
        lsttemp=[-40,70,20]
        #for tempiter in range(1,int(self.tempruns_D45)+1):
        for tempiter,actual_temp in enumerate(lsttemp):

            tempiter=tempiter+1
            self.check_temp_reached(actual_temp)
            #step 1 waiting for temperature ramp for 1 hr
            time.delay(60*60)
            #step 2 waiting at that temperature for 1 hr
            time.delay(60*60)

            #step 3 Acquistion by altering the position
            for cycle in range(1,7):
            # when ever  cycle is zero it is going to create a new file in autoitdatalogger45

                positionnumber =1
                self.desired_aerotech_pos(positionnumber)
                self.autoitdataloggerD45(self,tempiter,positionnumber,cycle)

                positionnumber =2
                self.desired_aerotech_pos(positionnumber)
                self.autoitdataloggerD45(self,tempiter,positionnumber,cycle)






    def completeacquisition(self):

        self.acq_starts_in_10sec()
        self.createdirectory(self.combolocation)
        #taking care of position homing and enable X , absolute
        self.aerotech_enab_absol()

        if (self.comboday_set == 'L1' or self.comboday_set == 'L2' or self.comboday_set == 'L3'):
            self.acquistionL123()

        if self.comboday_set == 'D4':
            self.acquistionD4()

        if self.comboday_set == 'D5':
            self.acquistionD5()

    def acq_starts_in_10sec(self):

        print("Acquisition starts in 10 seconds")
        for i in range(10,0,-1):
            print("i\n".format(i))
            time.sleep(1)





    #This function gets the data from the dataloger using autoit and using defaultlocation of the dataloger file
    #for 1 cycle. Then it iterates through it.
    def autoitdataloggerL123(self,tempiter,positionnumber):

        self.autoit_singlecycleclicks()
        files = []
        # r=root, d=directories, f = files
        for r, d, f in os.walk(self.path_to_datalogger_default):
            for file in f:
                files.append(os.path.join(r, file))


        with open(files[0],encoding='UTF-16') as f:
            df_datalogger_default = pd.read_csv(f,skiprows=20,header=None)
            if (positionnumber == 1 and tempiter == 1):
                df_datalogger_default.loc[[0]].to_csv(self.raw_text_path,index=False, header=False )  #heading
            df_datalogger_default.loc[[1]].to_csv(self.raw_text_path,index=False, header=False, mode='a')#This is where data is

        #combolocation folder is created when the acquisition is started for L123
        my_file=files[0]
        to_file=pathlib.Path(self.combolocation + tempiter + '_' + positionnumber +'.txt')
        shutil.move(my_file, to_file)


    #Changing the position to the desired value
    def desired_aerotech_pos(self,posnum):
        imme_comm=self.driver.find_element_by_id("immediate-command-text")
        time.sleep(0.5)
        imme_comm.clear()
        if posnum==1:
            imme_comm.send_keys("RAPID X -0.542000 F5")
            desired_pos=-0.542000
        elif posnum==2:
            imme_comm.send_keys("RAPID X 89.458000 F5")
            desired_pos=89.458000
        elif posnum==3:
            imme_comm.send_keys("RAPID X 179.45800 F5")
            desired_pos=179.45800
        elif posnum==4:
            imme_comm.send_keys("RAPID X 269.45800 F5")
            desired_pos=269.45800
        else:
            print("Invalid Position Number...")
            sys.exit(0)

        print("waiting for two minutes to set to desired position")
        time.sleep(1)
        imme_comm.send_keys(Keys.RETURN)
        time.sleep(60)
        self.driver.implicitly_wait(60)

        #Checking if the position feedback has reached the desired value
        pos_feedback_Text=0

        while abs(desired_pos-float(pos_feedback_Text))>10**-3:
            pos_feedback_Element=self.driver.find_element_by_id('axis0PosFbk')
            time.sleep(5)
            pos_feedback_Text = pos_feedback_Element.text

        print("Reached the desired position")
        return 1












    #This is the function that actually takes care of homing, enable,absolute in the aerotech position
    def aerotech_enab_absol(self):
        #Clicking two balls on the screen
        balls_Element=self.driver.find_element_by_id("enableDisableAxis0")

        if balls_Element is None:
            print("The position controller is not connecting...")
            sys.exit()
        else:
            print("Clicked Balls and waiting 60 seconds... ")
            balls_Element.click()
            time.sleep(1)
            self.driver.implicitly_wait(60)


        #Clicking home button
        print("Clicked home button and waiting 60 seconds... ")
        self.driver.find_element_by_id("homeAxis0").click()
        self.driver.implicitly_wait(60)



        #sending ENABLE X and pressing enter
        imme_comm=self.driver.find_element_by_id("immediate-command-text")
        imme_comm.send_keys("ENABLE X")
        time.sleep(1)
        imme_comm.send_keys(Keys.RETURN)
        time.sleep(5)


        #checking for NO ERROR in the  bottom status bar
        #bottombar_Element=driver.find_element_by_id("status-bar")
        #if bottombar_Element is not None:
        #    print("bottombar_Element found")
        #bottombar_value=bottombar_Element.get_attribute("value")
        #print("The bottombar value is " +bottombar_value)


        check_enable_Element = self.driver.find_element_by_id('axis0Status')
        check_enable_Text = check_enable_Element.text

        if check_enable_Text=='Enabled':
            print("Enabled and lets continue")
            time.sleep(0.5)

        #clearing the immediate-command text for next command and then Absolute command
        imme_comm=self.driver.find_element_by_id("immediate-command-text")
        time.sleep(0.5)
        imme_comm.clear()
        imme_comm.send_keys("ABSOLUTE")
        time.sleep(1)
        imme_comm.send_keys(Keys.RETURN)
        time.sleep(5)




    def autoit_singlecycleclicks(self):


        #Emptying all the files from the default directory of Datalogger saver so that only one file is created by datalogger3
        files = []
        # r=root, d=directories, f = files
        for r, d, f in os.walk(self.path_to_datalogger_default):
            for file in f:
                files.append(os.path.join(r, file))
        #print(files)

        for i in files:
            os.remove(i)

        print('Removed all files from datalogger default..starting acquisition')

        
        try:
            app = Application(backend="win32").connect(title=u'Configuration - 2 - BenchLink Data Logger 3', class_name='WindowsForms10.Window.8.app.0.33c0d9d')
            main_dlg = app[u'WindowsForms10.Window.8.app.0.33c0d9d']
            main_dlg.wait('visible')
            print("1")
            time.sleep(1)
            main_dlg.set_focus()
            print("2")
           
            p=main_dlg.TabControl.select(u'Scan and Log Data')
            #print(p.get_properties())
           
            p.client_rects()
           
            scanlog_dlg=main_dlg[u'WindowsForms10.Window.8.app.0.33c0d9d7']
            #scanlog_dlg.child_window(auto_id="m_gridInst", control_type="C1.Win.C1FlexGrid.C1FlexGrid").print_control_identifiers()
            scanlog_dlg.draw_outline()
            scanlog_dlg.click()
            time.sleep(0.5)
           
           
           
            send_keys('^{RIGHT}{DOWN}{LEFT}{LEFT}{ENTER}')
           
           
            app2 = Application().Connect(title=u'Set Data Log Fields', class_name='WindowsForms10.Window.8.app.0.33c0d9d')
            setdatalogfield_dlg = app2[u'WindowsForms10.Window.8.app.0.33c0d9d']
            print("11")
            setdatalogfield_dlg.wait('visible')
            print("22")
           
            checkboxbutton_notrequired = setdatalogfield_dlg.Button2
            checkboxbutton_notrequired.click()
           
            okbutton=setdatalogfield_dlg.OK
            okbutton.click()
           
           
            main_dlg.set_focus()
            print("2")
            send_keys('{RIGHT}{ENTER}')
           
        except(WindowNotFoundError):
            print ('window not found')
            pass
        except(WindowAmbiguousError):
            print ('There are too many  windows found')
            pass  
               
           
           
           
           
           
           
           
           
           
           
           
           
           
           
           
           
           
        #    send_keys("{VK_CONTROL down}"
        #          "{VK_RIGHT down}"
        #          "{VK_RIGHT up}"
        #          "{VK_CONTROL up}"
        #          "{VK_LEFT down}"
        #          "{VK_LEFT up}"
        #          "{VK_LEFT down}"
        #          "{VK_LEFT up}"
        #          "{ENTER}"
        #                             ) # to type PYWINAUTO
        
            #scanlog_dlg.print_control_identifiers()
            #app.WindowsForms10.Window.8.app.0.33c0d9d7.print_control_identifiers()
           
            #child_window(title="Scan and Log Data", auto_id="m_tpageScanLog", control_type="System.Windows.Forms.TabPage")
           
            #sys.stdout = open("scanandlog.txt","w")
            #scanlog_dlg.print_control_identifiers()
            #main_dlg.TabControl.Scan and Log Data.print_control_identifiers()  
            #app.Properties.child_window(title="Contains:", auto_id="13087", control_type="Edit")
            #main_dlg.TabControl.TabPage
           
        #    ctrl = app['WindowsForms10.Window.8.app.0.33c0d9d7']
        #    ctrl.Click()
        
           
        #    time.sleep(1)
        #    windowsformsbuttonappcdd = windowsformswindowappcdd.Button4
        #    time.sleep(1)
        #    windowsformsbuttonappcdd.click()
        #    time.sleep(1)
        #
        #
        #    app_dialog = app.top_window()
        #    time.sleep(1)
        #    app_dialog.set_focus()
        #    time.sleep(1)
        #    windowsformsbuttonappcdd = windowsformswindowappcdd.Cancel
        #    windowsformsbuttonappcdd.click()
        #    
        #    time.sleep(1)
        #    
        #    windowsformsbuttonappcdd2 = windowsformswindowappcdd.Button5
        #    windowsformsbuttonappcdd2.SetFocus()
        #    windowsformsbuttonappcdd2.select()
        
   
            #Click on the close button
            #end of the cycle



#if __name__ == '__main__':
root=TK.Tk()
window = Window(root)

print(window.intradelaydefault)
TK.mainloop()



#a = pd.DataFrame(data = [[1,2],[3,4]], index=range(2), columns = ['A', 'B'])
#c,b = a.iloc[0].copy(), a.iloc[1].copy()
#a.iloc[0],a.iloc[1] = b,c
#print ("Start : %s" % time.ctime())
